#!/usr/bin/env python3
"""
Create comprehensive Excel sheet with all STIs, questions, answers, gaps
Improved version that identifies gaps from answers themselves (e.g., "PIA not done")
"""

import pandas as pd
import sys
import argparse
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def extract_sti_from_text(text):
    """Extract STI codes from text (format: XXX-001)"""
    if pd.isna(text) or not text:
        return []
    text_str = str(text)
    # Pattern: 3-5 uppercase letters, dash, 1-3 digits
    pattern = r'\b([A-Z]{3,5}-\d{1,3})\b'
    matches = re.findall(pattern, text_str)
    return list(set(matches))  # Return unique matches

def identify_gaps_from_answer(answer_text, question_text, category, stis_in_answer=None):
    """Identify gaps from answer text itself and include STI names"""
    if pd.isna(answer_text) or not answer_text:
        return []
    
    answer_lower = str(answer_text).lower()
    question_lower = str(question_text).lower() if question_text else ''
    category_lower = str(category).lower() if category else ''
    stis_in_answer = stis_in_answer or []
    
    gaps = []
    
    # Gap indicators
    gap_keywords = [
        'not done', 'not completed', 'missing', 'no ', 'not available',
        'not implemented', 'not in place', 'lacking', 'absent',
        'cannot', 'unable', 'does not', "doesn't", 'failed',
        'incomplete', 'pending', 'waiting', 'not yet'
    ]
    
    # Check for gap indicators
    has_gap_indicator = any(keyword in answer_lower for keyword in gap_keywords)
    
    # Extract STIs mentioned in the answer text
    stis_mentioned = extract_sti_from_text(answer_text)
    all_stis_for_gap = list(set(stis_mentioned + stis_in_answer))
    
    # Build STI list for gap description
    sti_list_text = ''
    if all_stis_for_gap:
        if len(all_stis_for_gap) <= 10:
            sti_list_text = f" (STIs: {', '.join(all_stis_for_gap)})"
        else:
            sti_list_text = f" (STIs: {', '.join(all_stis_for_gap[:10])} and {len(all_stis_for_gap) - 10} more)"
    
    # Specific gap patterns
    gap_patterns = [
        (r'pia\s+(?:not\s+)?(?:done|completed)', 'PIA not completed'),
        (r'sia\s+(?:not\s+)?(?:done|completed)', 'SIA not completed'),
        (r'ess\s+(?:not\s+)?(?:done|completed)', 'ESS not completed'),
        (r'no\s+pia', 'PIA not done'),
        (r'no\s+sia', 'SIA not done'),
        (r'no\s+ess', 'ESS not done'),
        (r'missing\s+(?:response|documentation|evidence)', 'Missing response/documentation'),
        (r'cannot\s+verify', 'Cannot verify compliance'),
        (r'audit\s+trail\s+missing', 'Missing audit trail'),
        (r'no\s+monitoring', 'No monitoring/logging'),
    ]
    
    gap_name = ''
    gap_description = ''
    
    # Check for specific patterns
    for pattern, gap_name_pattern in gap_patterns:
        if re.search(pattern, answer_lower):
            gap_name = gap_name_pattern
            gap_description = f"{gap_name_pattern}{sti_list_text}: {str(answer_text)[:200]}"
            break
    
    # If no specific pattern but has gap indicators
    if not gap_name and has_gap_indicator:
        # Try to extract gap name from question or category
        if 'pia' in question_lower or 'pia' in category_lower:
            gap_name = 'PIA not completed'
        elif 'sia' in question_lower or 'sia' in category_lower:
            gap_name = 'SIA not completed'
        elif 'ess' in question_lower or 'ess' in category_lower:
            gap_name = 'ESS not completed'
        else:
            gap_name = 'Compliance gap identified'
        
        gap_description = f"Gap identified{sti_list_text}: {str(answer_text)[:300]}"
    
    # Check for explicit gap mentions
    if 'gap' in answer_lower:
        gap_match = re.search(r'gap[:\s]+([^\.\n]+)', answer_lower, re.IGNORECASE)
        if gap_match:
            gap_name = gap_match.group(1).strip()
            gap_description = f"{gap_name}{sti_list_text}: {str(answer_text)[:300]}"
    
    if gap_name:
        gaps.append({
            'Gap Name': gap_name,
            'Gap Description': gap_description,
            'Source': 'Answer Analysis',
            'STIs Affected': ', '.join(all_stis_for_gap) if all_stis_for_gap else ''
        })
    
    return gaps

def get_all_stis_from_data(df_master, df_gaps):
    """Extract all STIs from both Master and GAPS sheets"""
    all_stis = set()
    
    # Get STIs from GAPS sheet
    if 'STI' in df_gaps.columns:
        stis_from_gaps = df_gaps['STI'].dropna().unique()
        all_stis.update([str(sti).strip() for sti in stis_from_gaps if str(sti).strip() != 'N/A'])
    
    # Get STIs from Master sheet BU Response
    if 'BU Response ' in df_master.columns:
        for idx, row in df_master.iterrows():
            bu_response = row.get('BU Response ', '')
            stis = extract_sti_from_text(bu_response)
            all_stis.update(stis)
    
    # Get STIs from Master sheet BU Answer
    if 'BU Answer' in df_master.columns:
        for idx, row in df_master.iterrows():
            bu_answer = row.get('BU Answer', '')
            stis = extract_sti_from_text(bu_answer)
            all_stis.update(stis)
    
    return sorted(list(all_stis))

def create_comprehensive_sheet(excel_file, output_file):
    """Create comprehensive Excel with all STIs, questions, answers, gaps"""
    
    print("="*80)
    print("CREATING COMPREHENSIVE STI SHEET (V2 - Gap Detection from Answers)")
    print("="*80)
    print(f"\nReading Excel file: {excel_file}")
    
    try:
        # Read both sheets
        df_gaps = pd.read_excel(excel_file, sheet_name='GAPS')
        df_master = pd.read_excel(excel_file, sheet_name='Master')
    except FileNotFoundError:
        print(f"❌ Error: File not found: {excel_file}")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        sys.exit(1)
    
    print(f"✅ Loaded {len(df_gaps)} rows from GAPS sheet")
    print(f"✅ Loaded {len(df_master)} rows from Master sheet\n")
    
    # Get all STIs
    all_stis = get_all_stis_from_data(df_master, df_gaps)
    print(f"✅ Found {len(all_stis)} unique STIs\n")
    
    # Create a mapping of questions to gaps from GAPS sheet
    gap_lookup = {}  # {(STI, row_num): gap_info}
    
    for idx, gap_row in df_gaps.iterrows():
        sti = str(gap_row.get('STI', '')).strip() if pd.notna(gap_row.get('STI')) else ''
        if not sti or sti == 'N/A':
            continue
        
        row_ref = str(gap_row.get('Row Reference', '')).strip() if pd.notna(gap_row.get('Row Reference')) else ''
        row_num = None
        
        # Extract row number from reference
        if row_ref and 'Row' in row_ref:
            row_num_match = re.search(r'Row\s+(\d+)', row_ref)
            if row_num_match:
                row_num = int(row_num_match.group(1))
        
        gap_info = {
            'Gap Number': str(gap_row.get('Gap Number', '')).strip() if pd.notna(gap_row.get('Gap Number')) else '',
            'Gap Name': str(gap_row.get('AI Question', '')).strip() if pd.notna(gap_row.get('AI Question')) else '',
            'Gap Description': str(gap_row.get('Gap Description', '')).strip() if pd.notna(gap_row.get('Gap Description')) else '',
            'Domain': str(gap_row.get('Domain', '')).strip() if pd.notna(gap_row.get('Domain')) else '',
            'ESS Control': str(gap_row.get('ESS Control', '')).strip() if pd.notna(gap_row.get('ESS Control')) else '',
            'Source': 'GAPS Sheet'
        }
        
        # Store by STI and row number
        key = (sti, row_num)
        if key not in gap_lookup:
            gap_lookup[key] = []
        gap_lookup[key].append(gap_info)
    
    # Now create comprehensive rows for each STI and question
    comprehensive_data = []
    
    print("📊 Processing questions and answers for all STIs...")
    print("   Analyzing answers to identify gaps...\n")
    
    for master_idx, master_row in df_master.iterrows():
        # Get question
        question = str(master_row.get('Questions/Evidence Requirements', '')).strip() if pd.notna(master_row.get('Questions/Evidence Requirements')) else ''
        
        # Get answer
        bu_answer = str(master_row.get('BU Answer', '')).strip() if pd.notna(master_row.get('BU Answer')) else ''
        bu_response = str(master_row.get('BU Response ', '')).strip() if pd.notna(master_row.get('BU Response ')) else ''
        
        # Combine answers (prefer BU Answer, fallback to BU Response)
        answer = bu_answer if bu_answer else bu_response
        
        # Get other question details
        category = str(master_row.get('Category (Control)', '')).strip() if pd.notna(master_row.get('Category (Control)')) else ''
        hints = str(master_row.get('Hints/Guidance', '')).strip() if pd.notna(master_row.get('Hints/Guidance')) else ''
        
        # Get supporting evidence from multiple sources
        supporting_evidence = ''
        # Try BU Supporting Evidence first
        bu_supporting = str(master_row.get('BU Supporting Evidence', '')).strip() if pd.notna(master_row.get('BU Supporting Evidence')) else ''
        if bu_supporting and bu_supporting != 'nan' and len(bu_supporting) > 0:
            supporting_evidence = bu_supporting
        
        # Also check Supported Documents column
        if not supporting_evidence or len(supporting_evidence) < 5:
            supported_docs = str(master_row.get('Supported Documents - Good to Have for Readiness', '')).strip() if pd.notna(master_row.get('Supported Documents - Good to Have for Readiness')) else ''
            if supported_docs and supported_docs != 'nan' and len(supported_docs) > 0:
                if supporting_evidence:
                    supporting_evidence = f"{supporting_evidence}\n\nSupported Documents: {supported_docs}"
                else:
                    supporting_evidence = f"Supported Documents: {supported_docs}"
        
        # Also check if answer contains evidence/links
        if not supporting_evidence or len(supporting_evidence) < 5:
            # Look for URLs or links in the answer
            answer_text = answer if answer else bu_response
            if answer_text:
                # Look for various URL patterns
                url_patterns = [
                    r'https?://[^\s<>"\'\)]+',
                    r'confluence[^\s<>"\'\)]+',
                    r'google\.com/[^\s<>"\'\)]+',
                    r'drive\.google\.com/[^\s<>"\'\)]+',
                    r'redhat\.com/[^\s<>"\'\)]+',
                    r'github\.com/[^\s<>"\'\)]+',
                    r'gitlab\.com/[^\s<>"\'\)]+',
                    r'spaces\.redhat\.com/[^\s<>"\'\)]+',
                ]
                
                all_urls = []
                for pattern in url_patterns:
                    matches = re.findall(pattern, str(answer_text), re.IGNORECASE)
                    all_urls.extend(matches)
                
                # Remove duplicates while preserving order
                seen = set()
                unique_urls = []
                for url in all_urls:
                    if url not in seen:
                        seen.add(url)
                        unique_urls.append(url)
                
                if unique_urls:
                    # Format URLs nicely
                    if len(unique_urls) == 1:
                        supporting_evidence = f"Evidence: {unique_urls[0]}"
                    elif len(unique_urls) <= 5:
                        supporting_evidence = f"Evidence links: {', '.join(unique_urls)}"
                    else:
                        supporting_evidence = f"Evidence links: {', '.join(unique_urls[:5])} (and {len(unique_urls) - 5} more)"
                
                # Also look for document references in text
                if not supporting_evidence or len(supporting_evidence) < 10:
                    doc_keywords = ['document', 'link', 'reference', 'see', 'refer to', 'confluence', 'wiki']
                    answer_lower = str(answer_text).lower()
                    if any(keyword in answer_lower for keyword in doc_keywords):
                        # Extract sentences with document references
                        sentences = re.split(r'[.!?]\s+', str(answer_text))
                        doc_sentences = [s.strip() for s in sentences if any(kw in s.lower() for kw in doc_keywords)]
                        if doc_sentences:
                            if not supporting_evidence:
                                supporting_evidence = f"Document references: {doc_sentences[0][:200]}"
                            else:
                                supporting_evidence = f"{supporting_evidence}\n\nDocument references: {doc_sentences[0][:200]}"
        
        # Extract STIs mentioned in this question/answer
        stis_in_response = extract_sti_from_text(bu_response)
        stis_in_answer = extract_sti_from_text(bu_answer)
        stis_mentioned = list(set(stis_in_response + stis_in_answer))
        
        # If no STIs mentioned, check if this row has gaps from GAPS sheet
        if not stis_mentioned:
            row_num = master_idx + 1  # 1-indexed
            for sti in all_stis:
                key = (sti, row_num)
                if key in gap_lookup:
                    stis_mentioned.append(sti)
        
        # If still no STIs, check gaps sheet for this question
        if not stis_mentioned:
            matching_gaps = df_gaps[df_gaps['AI Question'].str.contains(question[:50], case=False, na=False, regex=False)]
            if len(matching_gaps) > 0:
                stis_from_gaps = matching_gaps['STI'].dropna().unique()
                stis_mentioned = [str(sti).strip() for sti in stis_from_gaps if str(sti).strip() != 'N/A']
        
        # Extract STIs from answer text first
        stis_in_answer_text = extract_sti_from_text(answer)
        all_stis_for_this_answer = list(set(stis_mentioned + stis_in_answer_text))
        
        # Analyze answer for gaps (pass STIs found in answer)
        gaps_from_answer = identify_gaps_from_answer(answer, question, category, all_stis_for_this_answer)
        
        # If answer indicates gaps but no specific STIs mentioned, use STIs from answer text
        if gaps_from_answer and not stis_mentioned:
            if stis_in_answer_text:
                stis_mentioned = stis_in_answer_text
        
        # Create entry for each STI mentioned
        if stis_mentioned:
            for sti in stis_mentioned:
                row_num = master_idx + 1
                key = (sti, row_num)
                gaps_from_sheet = gap_lookup.get(key, [])
                
                # Combine gaps from sheet and gaps from answer analysis
                all_gaps = gaps_from_sheet.copy()
                
                # Add gaps identified from answer
                for gap_from_answer in gaps_from_answer:
                    # Check if this gap is already in the list
                    if not any(g['Gap Description'] == gap_from_answer['Gap Description'] for g in all_gaps):
                        all_gaps.append(gap_from_answer)
                
                if all_gaps:
                    # Create entry for each gap
                    for gap in all_gaps:
                        # Enhance gap description with STI list if not already included
                        gap_desc = gap.get('Gap Description', '')
                        stis_affected = gap.get('STIs Affected', '')
                        
                        # If gap description doesn't include STIs and we have them, add them
                        if stis_affected and sti not in gap_desc:
                            if gap_desc and not gap_desc.endswith(')'):
                                gap_desc = f"{gap_desc} (STIs: {sti}"
                                if stis_affected and stis_affected != sti:
                                    gap_desc = f"{gap_desc}, {stis_affected}"
                                gap_desc = f"{gap_desc})"
                            elif not gap_desc:
                                gap_desc = f"STIs affected: {sti}"
                                if stis_affected and stis_affected != sti:
                                    gap_desc = f"{gap_desc}, {stis_affected}"
                        
                        comprehensive_data.append({
                            'STI': sti,
                            'Question': question,
                            'Answer': answer,
                            'Gap Name': gap.get('Gap Name', ''),
                            'Gap Description': gap_desc,
                            'Gap Number': gap.get('Gap Number', ''),
                            'Domain': gap.get('Domain', ''),
                            'ESS Control': gap.get('ESS Control', ''),
                            'Category': category,
                            'Hints/Guidance': hints,
                            'Supporting Evidence': supporting_evidence,
                            'Row Number': row_num,
                            'Has Gap': 'Yes',
                            'Gap Source': gap.get('Source', 'Unknown'),
                            'STIs Affected': stis_affected if stis_affected else sti
                        })
                else:
                    # No gap for this STI and question
                    comprehensive_data.append({
                        'STI': sti,
                        'Question': question,
                        'Answer': answer,
                        'Gap Name': '',
                        'Gap Description': 'No gap identified',
                        'Gap Number': '',
                        'Domain': '',
                        'ESS Control': '',
                        'Category': category,
                        'Hints/Guidance': hints,
                        'Supporting Evidence': supporting_evidence,
                        'Row Number': row_num,
                        'Has Gap': 'No',
                        'Gap Source': '',
                        'STIs Affected': ''
                    })
        else:
            # No STIs mentioned but answer might indicate gaps
            # Check if answer indicates gaps for all STIs or specific ones
            if gaps_from_answer:
                # Answer indicates gaps but no specific STIs
                # This might be a general gap - we'll handle it separately
                pass
    
    # Also add entries for STIs that have gaps but weren't in master responses
    print("📊 Adding gaps from GAPS sheet that weren't linked to questions...")
    for idx, gap_row in df_gaps.iterrows():
        sti = str(gap_row.get('STI', '')).strip() if pd.notna(gap_row.get('STI')) else ''
        if not sti or sti == 'N/A':
            continue
        
        gap_name = str(gap_row.get('AI Question', '')).strip() if pd.notna(gap_row.get('AI Question')) else ''
        gap_desc = str(gap_row.get('Gap Description', '')).strip() if pd.notna(gap_row.get('Gap Description')) else ''
        
        # Check if we already have this gap in comprehensive_data
        already_added = any(
            d['STI'] == sti and 
            d['Gap Name'] == gap_name and 
            d['Gap Description'] == gap_desc
            for d in comprehensive_data
        )
        
        if not already_added:
            # Enhance gap description with STI name
            gap_desc_enhanced = gap_desc
            if sti and sti not in gap_desc:
                gap_desc_enhanced = f"{gap_desc} (STI: {sti})"
            
            # Try to find supporting evidence from Master sheet if we have row reference
            supporting_evidence = ''
            row_ref = str(gap_row.get('Row Reference', '')).strip() if pd.notna(gap_row.get('Row Reference')) else ''
            if row_ref and 'Row' in row_ref:
                try:
                    row_num_match = re.search(r'Row\s+(\d+)', row_ref)
                    if row_num_match:
                        master_row_num = int(row_num_match.group(1)) - 1  # Convert to 0-indexed
                        if 0 <= master_row_num < len(df_master):
                            master_row = df_master.iloc[master_row_num]
                            # Get supporting evidence
                            bu_supporting = str(master_row.get('BU Supporting Evidence', '')).strip() if pd.notna(master_row.get('BU Supporting Evidence')) else ''
                            if bu_supporting and bu_supporting != 'nan':
                                supporting_evidence = bu_supporting
                            
                            # Also check Supported Documents
                            if not supporting_evidence or len(supporting_evidence) < 5:
                                supported_docs = str(master_row.get('Supported Documents - Good to Have for Readiness', '')).strip() if pd.notna(master_row.get('Supported Documents - Good to Have for Readiness')) else ''
                                if supported_docs and supported_docs != 'nan':
                                    supporting_evidence = f"Supported Documents: {supported_docs}"
                except:
                    pass
            
            comprehensive_data.append({
                'STI': sti,
                'Question': gap_name,  # Use gap name as question if no question found
                'Answer': '',
                'Gap Name': gap_name,
                'Gap Description': gap_desc_enhanced,
                'Gap Number': str(gap_row.get('Gap Number', '')).strip() if pd.notna(gap_row.get('Gap Number')) else '',
                'Domain': str(gap_row.get('Domain', '')).strip() if pd.notna(gap_row.get('Domain')) else '',
                'ESS Control': str(gap_row.get('ESS Control', '')).strip() if pd.notna(gap_row.get('ESS Control')) else '',
                'Category': '',
                'Hints/Guidance': '',
                'Supporting Evidence': supporting_evidence,
                'Row Number': '',
                'Has Gap': 'Yes',
                'Gap Source': 'GAPS Sheet',
                'STIs Affected': sti
            })
    
    # Ensure all STIs are represented
    print("📊 Ensuring all STIs are represented...")
    stis_in_data = set([d['STI'] for d in comprehensive_data])
    missing_stis = [sti for sti in all_stis if sti not in stis_in_data]
    
    for sti in missing_stis:
        comprehensive_data.append({
            'STI': sti,
            'Question': 'No questions/assessments found',
            'Answer': '',
            'Gap Name': '',
            'Gap Description': 'No data available for this STI',
            'Gap Number': '',
            'Domain': '',
            'ESS Control': '',
            'Category': '',
            'Hints/Guidance': '',
            'Supporting Evidence': '',
            'Row Number': '',
            'Has Gap': 'No Data',
            'Gap Source': '',
            'STIs Affected': ''
        })
    
    print(f"✅ Created {len(comprehensive_data)} comprehensive entries\n")
    
    # Create DataFrame
    df_comprehensive = pd.DataFrame(comprehensive_data)
    
    # Sort by STI, then by Has Gap (Yes first), then by Question
    df_comprehensive = df_comprehensive.sort_values(['STI', 'Has Gap', 'Question'], 
                                                     ascending=[True, False, True],
                                                     na_position='last')
    
    # Create Excel file with formatting
    print("📝 Creating Excel file...")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Main comprehensive sheet
        df_comprehensive.to_excel(writer, sheet_name='All STIs - Questions & Gaps', index=False)
        
        # Create summary sheet
        summary_data = []
        for sti in all_stis:
            sti_rows = df_comprehensive[df_comprehensive['STI'] == sti]
            gaps = sti_rows[sti_rows['Has Gap'] == 'Yes']
            no_gaps = sti_rows[sti_rows['Has Gap'] == 'No']
            no_data = sti_rows[sti_rows['Has Gap'] == 'No Data']
            
            summary_data.append({
                'STI': sti,
                'Total Questions': len(sti_rows),
                'Questions with Gaps': len(gaps),
                'Questions without Gaps': len(no_gaps),
                'No Data': len(no_data),
                'Domains Affected': ', '.join(gaps['Domain'].dropna().unique().tolist()) if len(gaps) > 0 else 'None',
                'ESS Controls Affected': ', '.join(gaps['ESS Control'].dropna().unique().tolist()) if len(gaps) > 0 else 'None',
                'Status': 'Has Gaps' if len(gaps) > 0 else ('No Data' if len(no_data) > 0 else 'No Gaps')
            })
        
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='STI Summary', index=False)
        
        # Format both sheets
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            
            # Style the header row
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
            
            # Auto-adjust column widths
            df_to_use = df_comprehensive if sheet_name == 'All STIs - Questions & Gaps' else df_summary
            for col_idx, col in enumerate(df_to_use.columns, 1):
                max_length = len(str(col))
                col_letter = get_column_letter(col_idx)
                
                # Check data rows
                for row in worksheet.iter_rows(min_row=2, max_row=min(worksheet.max_row, 1000)):  # Limit check for performance
                    if row[col_idx - 1].value:
                        cell_value = str(row[col_idx - 1].value)
                        max_length = max(max_length, len(cell_value))
                
                adjusted_width = min(max_length + 2, 150)  # Cap at 150
                worksheet.column_dimensions[col_letter].width = adjusted_width
            
            # Wrap text in cells and add borders
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    cell.border = border
            
            # Freeze header row
            worksheet.freeze_panes = 'A2'
            
            # Add conditional formatting for gaps
            if sheet_name == 'All STIs - Questions & Gaps':
                gap_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                no_gap_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
                
                # Color rows with gaps
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    has_gap_cell = row[12]  # 'Has Gap' column (M)
                    if has_gap_cell.value == 'Yes':
                        for cell in row:
                            cell.fill = gap_fill
                    elif has_gap_cell.value == 'No':
                        for cell in row:
                            cell.fill = no_gap_fill
    
    print(f"\n✅ Excel file created: {output_file}")
    print(f"   Sheet 1: All STIs - Questions & Gaps - {len(df_comprehensive)} rows")
    print(f"   Sheet 2: STI Summary - {len(df_summary)} STIs")
    print(f"\n📊 Statistics:")
    print(f"   Total STIs: {len(all_stis)}")
    print(f"   STIs with gaps: {len(df_summary[df_summary['Status'] == 'Has Gaps'])}")
    print(f"   STIs without gaps: {len(df_summary[df_summary['Status'] == 'No Gaps'])}")
    print(f"   STIs with no data: {len(df_summary[df_summary['Status'] == 'No Data'])}")
    print(f"   Total entries: {len(df_comprehensive)}")
    print(f"   Entries with gaps: {len(df_comprehensive[df_comprehensive['Has Gap'] == 'Yes'])}")
    print(f"   Entries without gaps: {len(df_comprehensive[df_comprehensive['Has Gap'] == 'No'])}")
    
    # Show sample gaps identified from answers
    answer_gaps = df_comprehensive[df_comprehensive['Gap Source'] == 'Answer Analysis']
    if len(answer_gaps) > 0:
        print(f"\n🔍 Gaps identified from answer analysis: {len(answer_gaps)}")
        print("   Sample gaps found in answers:")
        for idx, row in answer_gaps.head(5).iterrows():
            print(f"      - {row['STI']}: {row['Gap Name']}")

def main():
    parser = argparse.ArgumentParser(description="Create comprehensive Excel sheet with all STIs, questions, answers, and gaps (V2 - detects gaps from answers)")
    parser.add_argument("--input", 
                       default="/Users/kwhelan/Downloads/AU2 Secure Flow + SSC - Simplified Questions - Internal Audit Prep - Template (4).xlsx",
                       help="Input Excel file path")
    parser.add_argument("--output", 
                       default="Secure_Flow_All_STIs_Comprehensive.xlsx",
                       help="Output Excel file name")
    args = parser.parse_args()
    
    create_comprehensive_sheet(args.input, args.output)

if __name__ == "__main__":
    main()

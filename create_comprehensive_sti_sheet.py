#!/usr/bin/env python3
"""
Create comprehensive Excel sheet with all STIs, questions, answers, gaps
For every STI, shows all questions, answers, and gap information
"""

import pandas as pd
import sys
import argparse
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def extract_sti_from_text(text):
    """Extract STI codes from text (format: XXX-001)"""
    if pd.isna(text) or not text:
        return []
    text_str = str(text)
    # Pattern: 3-5 uppercase letters, dash, 1-3 digits
    pattern = r'\b([A-Z]{3,5}-\d{1,3})\b'
    matches = re.findall(pattern, text_str)
    return list(set(matches))  # Return unique matches

def get_all_stis_from_data(df_master, df_gaps):
    """Extract all STIs from both Master and GAPS sheets"""
    all_stis = set()
    
    # Get STIs from GAPS sheet
    if 'STI' in df_gaps.columns:
        stis_from_gaps = df_gaps['STI'].dropna().unique()
        all_stis.update([str(sti).strip() for sti in stis_from_gaps if str(sti).strip() != 'N/A'])
    
    # Get STIs from Master sheet BU Response
    if 'BU Response ' in df_master.columns:
        for idx, row in df_master.iterrows():
            bu_response = row.get('BU Response ', '')
            stis = extract_sti_from_text(bu_response)
            all_stis.update(stis)
    
    # Get STIs from Master sheet BU Answer
    if 'BU Answer' in df_master.columns:
        for idx, row in df_master.iterrows():
            bu_answer = row.get('BU Answer', '')
            stis = extract_sti_from_text(bu_answer)
            all_stis.update(stis)
    
    return sorted(list(all_stis))

def create_comprehensive_sheet(excel_file, output_file):
    """Create comprehensive Excel with all STIs, questions, answers, gaps"""
    
    print("="*80)
    print("CREATING COMPREHENSIVE STI SHEET")
    print("="*80)
    print(f"\nReading Excel file: {excel_file}")
    
    try:
        # Read both sheets
        df_gaps = pd.read_excel(excel_file, sheet_name='GAPS')
        df_master = pd.read_excel(excel_file, sheet_name='Master')
    except FileNotFoundError:
        print(f"❌ Error: File not found: {excel_file}")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        sys.exit(1)
    
    print(f"✅ Loaded {len(df_gaps)} rows from GAPS sheet")
    print(f"✅ Loaded {len(df_master)} rows from Master sheet\n")
    
    # Get all STIs
    all_stis = get_all_stis_from_data(df_master, df_gaps)
    print(f"✅ Found {len(all_stis)} unique STIs\n")
    
    # Create a mapping of questions to gaps
    # First, process gaps to create gap lookup
    gap_lookup = {}  # {(STI, row_num): gap_info}
    
    for idx, gap_row in df_gaps.iterrows():
        sti = str(gap_row.get('STI', '')).strip() if pd.notna(gap_row.get('STI')) else ''
        if not sti or sti == 'N/A':
            continue
        
        row_ref = str(gap_row.get('Row Reference', '')).strip() if pd.notna(gap_row.get('Row Reference')) else ''
        row_num = None
        
        # Extract row number from reference
        if row_ref and 'Row' in row_ref:
            row_num_match = re.search(r'Row\s+(\d+)', row_ref)
            if row_num_match:
                row_num = int(row_num_match.group(1))
        
        gap_info = {
            'Gap Number': str(gap_row.get('Gap Number', '')).strip() if pd.notna(gap_row.get('Gap Number')) else '',
            'Gap Name': str(gap_row.get('AI Question', '')).strip() if pd.notna(gap_row.get('AI Question')) else '',
            'Gap Description': str(gap_row.get('Gap Description', '')).strip() if pd.notna(gap_row.get('Gap Description')) else '',
            'Domain': str(gap_row.get('Domain', '')).strip() if pd.notna(gap_row.get('Domain')) else '',
            'ESS Control': str(gap_row.get('ESS Control', '')).strip() if pd.notna(gap_row.get('ESS Control')) else '',
        }
        
        # Store by STI and row number
        key = (sti, row_num)
        if key not in gap_lookup:
            gap_lookup[key] = []
        gap_lookup[key].append(gap_info)
    
    # Now create comprehensive rows for each STI and question
    comprehensive_data = []
    
    print("📊 Processing questions and answers for all STIs...")
    
    for master_idx, master_row in df_master.iterrows():
        # Get question
        question = str(master_row.get('Questions/Evidence Requirements', '')).strip() if pd.notna(master_row.get('Questions/Evidence Requirements')) else ''
        
        # Get answer
        bu_answer = str(master_row.get('BU Answer', '')).strip() if pd.notna(master_row.get('BU Answer')) else ''
        bu_response = str(master_row.get('BU Response ', '')).strip() if pd.notna(master_row.get('BU Response ')) else ''
        
        # Combine answers
        answer = bu_answer if bu_answer else bu_response
        
        # Get other question details
        category = str(master_row.get('Category (Control)', '')).strip() if pd.notna(master_row.get('Category (Control)')) else ''
        hints = str(master_row.get('Hints/Guidance', '')).strip() if pd.notna(master_row.get('Hints/Guidance')) else ''
        supporting_evidence = str(master_row.get('BU Supporting Evidence', '')).strip() if pd.notna(master_row.get('BU Supporting Evidence')) else ''
        
        # Extract STIs mentioned in this question/answer
        stis_in_response = extract_sti_from_text(bu_response)
        stis_in_answer = extract_sti_from_text(bu_answer)
        stis_mentioned = list(set(stis_in_response + stis_in_answer))
        
        # If no STIs mentioned, check if this row has gaps
        if not stis_mentioned:
            # Check if any gap references this row
            row_num = master_idx + 1  # 1-indexed
            for sti in all_stis:
                key = (sti, row_num)
                if key in gap_lookup:
                    stis_mentioned.append(sti)
        
        # If still no STIs, create entry for "All STIs" or check gaps sheet
        if not stis_mentioned:
            # Check gaps sheet for this question
            matching_gaps = df_gaps[df_gaps['AI Question'].str.contains(question[:50], case=False, na=False, regex=False)]
            if len(matching_gaps) > 0:
                stis_from_gaps = matching_gaps['STI'].dropna().unique()
                stis_mentioned = [str(sti).strip() for sti in stis_from_gaps if str(sti).strip() != 'N/A']
        
        # Create entry for each STI mentioned
        if stis_mentioned:
            for sti in stis_mentioned:
                # Check for gaps for this STI and row
                row_num = master_idx + 1
                key = (sti, row_num)
                gaps_for_this = gap_lookup.get(key, [])
                
                if gaps_for_this:
                    # Create entry for each gap
                    for gap in gaps_for_this:
                        comprehensive_data.append({
                            'STI': sti,
                            'Question': question,
                            'Answer': answer,
                            'Gap Name': gap['Gap Name'],
                            'Gap Description': gap['Gap Description'],
                            'Gap Number': gap['Gap Number'],
                            'Domain': gap['Domain'],
                            'ESS Control': gap['ESS Control'],
                            'Category': category,
                            'Hints/Guidance': hints,
                            'Supporting Evidence': supporting_evidence,
                            'Row Number': row_num,
                            'Has Gap': 'Yes'
                        })
                else:
                    # No gap for this STI and question
                    comprehensive_data.append({
                        'STI': sti,
                        'Question': question,
                        'Answer': answer,
                        'Gap Name': '',
                        'Gap Description': 'No gap identified',
                        'Gap Number': '',
                        'Domain': '',
                        'ESS Control': '',
                        'Category': category,
                        'Hints/Guidance': hints,
                        'Supporting Evidence': supporting_evidence,
                        'Row Number': row_num,
                        'Has Gap': 'No'
                    })
        else:
            # No STIs mentioned - create entry for all STIs or mark as general
            # Actually, skip entries with no STI association
            pass
    
    # Also add entries for STIs that have gaps but weren't in master responses
    print("📊 Adding gaps that weren't linked to questions...")
    for idx, gap_row in df_gaps.iterrows():
        sti = str(gap_row.get('STI', '')).strip() if pd.notna(gap_row.get('STI')) else ''
        if not sti or sti == 'N/A':
            continue
        
        gap_name = str(gap_row.get('AI Question', '')).strip() if pd.notna(gap_row.get('AI Question')) else ''
        gap_desc = str(gap_row.get('Gap Description', '')).strip() if pd.notna(gap_row.get('Gap Description')) else ''
        
        # Check if we already have this gap in comprehensive_data
        already_added = any(
            d['STI'] == sti and 
            d['Gap Name'] == gap_name and 
            d['Gap Description'] == gap_desc
            for d in comprehensive_data
        )
        
        if not already_added:
            comprehensive_data.append({
                'STI': sti,
                'Question': gap_name,  # Use gap name as question if no question found
                'Answer': '',
                'Gap Name': gap_name,
                'Gap Description': gap_desc,
                'Gap Number': str(gap_row.get('Gap Number', '')).strip() if pd.notna(gap_row.get('Gap Number')) else '',
                'Domain': str(gap_row.get('Domain', '')).strip() if pd.notna(gap_row.get('Domain')) else '',
                'ESS Control': str(gap_row.get('ESS Control', '')).strip() if pd.notna(gap_row.get('ESS Control')) else '',
                'Category': '',
                'Hints/Guidance': '',
                'Supporting Evidence': '',
                'Row Number': '',
                'Has Gap': 'Yes'
            })
    
    # Ensure all STIs are represented
    print("📊 Ensuring all STIs are represented...")
    stis_in_data = set([d['STI'] for d in comprehensive_data])
    missing_stis = [sti for sti in all_stis if sti not in stis_in_data]
    
    for sti in missing_stis:
        comprehensive_data.append({
            'STI': sti,
            'Question': 'No questions/assessments found',
            'Answer': '',
            'Gap Name': '',
            'Gap Description': 'No data available for this STI',
            'Gap Number': '',
            'Domain': '',
            'ESS Control': '',
            'Category': '',
            'Hints/Guidance': '',
            'Supporting Evidence': '',
            'Row Number': '',
            'Has Gap': 'No Data'
        })
    
    print(f"✅ Created {len(comprehensive_data)} comprehensive entries\n")
    
    # Create DataFrame
    df_comprehensive = pd.DataFrame(comprehensive_data)
    
    # Sort by STI, then by Has Gap (Yes first), then by Question
    df_comprehensive = df_comprehensive.sort_values(['STI', 'Has Gap', 'Question'], 
                                                     ascending=[True, False, True],
                                                     na_position='last')
    
    # Create Excel file with formatting
    print("📝 Creating Excel file...")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Main comprehensive sheet
        df_comprehensive.to_excel(writer, sheet_name='All STIs - Questions & Gaps', index=False)
        
        # Create summary sheet
        summary_data = []
        for sti in all_stis:
            sti_rows = df_comprehensive[df_comprehensive['STI'] == sti]
            gaps = sti_rows[sti_rows['Has Gap'] == 'Yes']
            no_gaps = sti_rows[sti_rows['Has Gap'] == 'No']
            no_data = sti_rows[sti_rows['Has Gap'] == 'No Data']
            
            summary_data.append({
                'STI': sti,
                'Total Questions': len(sti_rows),
                'Questions with Gaps': len(gaps),
                'Questions without Gaps': len(no_gaps),
                'No Data': len(no_data),
                'Domains Affected': ', '.join(gaps['Domain'].dropna().unique().tolist()) if len(gaps) > 0 else 'None',
                'ESS Controls Affected': ', '.join(gaps['ESS Control'].dropna().unique().tolist()) if len(gaps) > 0 else 'None',
                'Status': 'Has Gaps' if len(gaps) > 0 else ('No Data' if len(no_data) > 0 else 'No Gaps')
            })
        
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='STI Summary', index=False)
        
        # Format both sheets
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            
            # Style the header row
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
            
            # Auto-adjust column widths
            df_to_use = df_comprehensive if sheet_name == 'All STIs - Questions & Gaps' else df_summary
            for col_idx, col in enumerate(df_to_use.columns, 1):
                max_length = len(str(col))
                col_letter = get_column_letter(col_idx)
                
                # Check data rows
                for row in worksheet.iter_rows(min_row=2, max_row=min(worksheet.max_row, 1000)):  # Limit check for performance
                    if row[col_idx - 1].value:
                        cell_value = str(row[col_idx - 1].value)
                        max_length = max(max_length, len(cell_value))
                
                adjusted_width = min(max_length + 2, 150)  # Cap at 150
                worksheet.column_dimensions[col_letter].width = adjusted_width
            
            # Wrap text in cells and add borders
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    cell.border = border
            
            # Freeze header row
            worksheet.freeze_panes = 'A2'
            
            # Add conditional formatting for gaps
            if sheet_name == 'All STIs - Questions & Gaps':
                from openpyxl.formatting.rule import CellIsRule
                gap_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                no_gap_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
                
                # Color rows with gaps
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    has_gap_cell = row[12]  # 'Has Gap' column (M)
                    if has_gap_cell.value == 'Yes':
                        for cell in row:
                            cell.fill = gap_fill
                    elif has_gap_cell.value == 'No':
                        for cell in row:
                            cell.fill = no_gap_fill
    
    print(f"\n✅ Excel file created: {output_file}")
    print(f"   Sheet 1: All STIs - Questions & Gaps - {len(df_comprehensive)} rows")
    print(f"   Sheet 2: STI Summary - {len(df_summary)} STIs")
    print(f"\n📊 Statistics:")
    print(f"   Total STIs: {len(all_stis)}")
    print(f"   STIs with gaps: {len(df_summary[df_summary['Status'] == 'Has Gaps'])}")
    print(f"   STIs without gaps: {len(df_summary[df_summary['Status'] == 'No Gaps'])}")
    print(f"   STIs with no data: {len(df_summary[df_summary['Status'] == 'No Data'])}")
    print(f"   Total entries: {len(df_comprehensive)}")
    print(f"   Entries with gaps: {len(df_comprehensive[df_comprehensive['Has Gap'] == 'Yes'])}")
    print(f"   Entries without gaps: {len(df_comprehensive[df_comprehensive['Has Gap'] == 'No'])}")

def main():
    parser = argparse.ArgumentParser(description="Create comprehensive Excel sheet with all STIs, questions, answers, and gaps")
    parser.add_argument("--input", 
                       default="/Users/kwhelan/Downloads/AU2 Secure Flow + SSC - Simplified Questions - Internal Audit Prep - Template (4).xlsx",
                       help="Input Excel file path")
    parser.add_argument("--output", 
                       default="Secure_Flow_All_STIs_Comprehensive.xlsx",
                       help="Output Excel file name")
    args = parser.parse_args()
    
    create_comprehensive_sheet(args.input, args.output)

if __name__ == "__main__":
    main()

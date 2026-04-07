#!/usr/bin/env python3
"""
Extract all Secure Flow SP compliance gaps from Excel file
Creates an Excel sheet with all gaps (excluding missing responses/documentation)
"""

import pandas as pd
import sys
import argparse
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def extract_gaps(excel_file, output_file):
    """Extract compliance gaps from GAPS sheet"""
    
    print(f"Reading Excel file: {excel_file}")
    try:
        # Read the GAPS sheet
        df = pd.read_excel(excel_file, sheet_name='GAPS')
    except FileNotFoundError:
        print(f"❌ Error: File not found: {excel_file}")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        print("\nTrying to list available sheets...")
        try:
            xl = pd.ExcelFile(excel_file)
            print(f"Available sheets: {xl.sheet_names}")
        except:
            pass
        sys.exit(1)
    
    print(f"✅ Loaded {len(df)} rows from GAPS sheet\n")
    print(f"Columns: {list(df.columns)}\n")
    
    # Identify key columns - GAPS sheet already has extracted gaps
    gap_desc_col = None
    question_col = None
    sti_col = None
    domain_col = None
    ess_control_col = None
    gap_num_col = None
    
    for col in df.columns:
        col_lower = str(col).lower()
        if 'gap description' in col_lower:
            gap_desc_col = col
        elif 'ai question' in col_lower or ('question' in col_lower and 'evidence' in col_lower):
            question_col = col
        elif 'sti' in col_lower:
            sti_col = col
        elif 'domain' in col_lower:
            domain_col = col
        elif 'ess' in col_lower and 'control' in col_lower:
            ess_control_col = col
        elif 'gap number' in col_lower:
            gap_num_col = col
    
    print("Identified columns:")
    print(f"  Gap Description: {gap_desc_col}")
    print(f"  AI Question: {question_col}")
    print(f"  STI: {sti_col}")
    print(f"  Domain: {domain_col}")
    print(f"  ESS Control: {ess_control_col}")
    print(f"  Gap Number: {gap_num_col}\n")
    
    if not gap_desc_col:
        print("❌ Could not find 'Gap Description' column")
        print("Available columns:")
        for col in df.columns:
            print(f"  - {col}")
        sys.exit(1)
    
    # Filter for gaps (non-empty gap descriptions)
    # Exclude "missing responses" and "missing documentation"
    gaps = []
    
    for idx, row in df.iterrows():
        gap_desc = str(row.get(gap_desc_col, '')).strip() if pd.notna(row.get(gap_desc_col)) else ''
        
        # Skip empty gap descriptions
        if not gap_desc or gap_desc.lower() in ['nan', 'none', '']:
            continue
        
        # Exclude "missing responses" and "missing documentation"
        gap_desc_lower = gap_desc.lower()
        if 'missing response' in gap_desc_lower or 'missing documentation' in gap_desc_lower:
            continue
        
        # This is a gap
        gap_data = {
            'Gap Number': str(row.get(gap_num_col, '')).strip() if gap_num_col and pd.notna(row.get(gap_num_col)) else idx + 1,
            'STI': str(row.get(sti_col, 'N/A')).strip() if sti_col and pd.notna(row.get(sti_col)) else 'N/A',
            'Domain': str(row.get(domain_col, 'N/A')).strip() if domain_col and pd.notna(row.get(domain_col)) else 'N/A',
            'ESS Control': str(row.get(ess_control_col, 'N/A')).strip() if ess_control_col and pd.notna(row.get(ess_control_col)) else 'N/A',
            'AI Question': str(row.get(question_col, 'N/A')).strip() if question_col and pd.notna(row.get(question_col)) else 'N/A',
            'Gap Description': gap_desc,
            'Row Reference': str(row.get('Row Reference', '')).strip() if pd.notna(row.get('Row Reference')) else ''
        }
        
        gaps.append(gap_data)
    
    print(f"✅ Found {len(gaps)} compliance gaps (excluding missing responses/documentation)\n")
    
    if not gaps:
        print("⚠️  No gaps found. Check the BU Response column for gap descriptions.")
        return
    
    # Create DataFrame
    df_gaps = pd.DataFrame(gaps)
    
    # Create Excel file with formatting
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_gaps.to_excel(writer, sheet_name='All Gaps', index=False)
        
        # Get the worksheet
        worksheet = writer.sheets['All Gaps']
        
        # Style the header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Auto-adjust column widths
        for col_idx, col in enumerate(df_gaps.columns, 1):
            max_length = max(
                df_gaps[col].astype(str).map(len).max(),
                len(str(col))
            )
            adjusted_width = min(max_length + 2, 100)  # Cap at 100
            worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
        
        # Wrap text in cells
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        # Freeze header row
        worksheet.freeze_panes = 'A2'
    
    print(f"✅ Excel file created: {output_file}")
    print(f"   Total gaps: {len(gaps)}")
    print(f"\nGap breakdown by Domain:")
    if domain_col:
        domain_counts = df_gaps['Domain'].value_counts()
        for domain, count in domain_counts.items():
            print(f"  {domain}: {count}")
    
    print(f"\nGap breakdown by STI:")
    sti_counts = df_gaps['STI'].value_counts()
    for sti, count in sti_counts.head(10).items():
        print(f"  {sti}: {count}")
    if len(sti_counts) > 10:
        print(f"  ... and {len(sti_counts) - 10} more STIs")

def main():
    parser = argparse.ArgumentParser(description="Extract all Secure Flow SP compliance gaps to Excel")
    parser.add_argument("--input", 
                       default="/Users/kwhelan/Downloads/AU2 Secure Flow + SSC - Simplified Questions - Internal Audit Prep - Template (4).xlsx",
                       help="Input Excel file path")
    parser.add_argument("--output", 
                       default="Secure_Flow_SP_Gaps.xlsx",
                       help="Output Excel file name")
    args = parser.parse_args()
    
    extract_gaps(args.input, args.output)

if __name__ == "__main__":
    main()


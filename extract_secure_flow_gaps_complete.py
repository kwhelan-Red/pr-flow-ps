#!/usr/bin/env python3
"""
Extract all Secure Flow SP compliance gaps with full details
Includes all STIs and pulls detailed comments from Master sheet
"""

import pandas as pd
import sys
import argparse
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def extract_sti_from_text(text):
    """Extract STI codes from text (format: XXX-001)"""
    if pd.isna(text) or not text:
        return []
    text_str = str(text)
    # Pattern: 3-5 uppercase letters, dash, 1-3 digits
    pattern = r'\b([A-Z]{3,5}-\d{1,3})\b'
    matches = re.findall(pattern, text_str)
    return list(set(matches))  # Return unique matches

def get_all_stis_from_master(df_master):
    """Extract all STIs mentioned in Master sheet"""
    all_stis = set()
    
    # Check BU Response column for STI references
    if 'BU Response ' in df_master.columns:
        for idx, row in df_master.iterrows():
            bu_response = row.get('BU Response ', '')
            stis = extract_sti_from_text(bu_response)
            all_stis.update(stis)
    
    # Check BU Answer column
    if 'BU Answer' in df_master.columns:
        for idx, row in df_master.iterrows():
            bu_answer = row.get('BU Answer', '')
            stis = extract_sti_from_text(bu_answer)
            all_stis.update(stis)
    
    return sorted(list(all_stis))

def extract_gaps_with_details(excel_file, output_file):
    """Extract compliance gaps with full details from Master sheet"""
    
    print(f"Reading Excel file: {excel_file}")
    try:
        # Read both sheets
        df_gaps = pd.read_excel(excel_file, sheet_name='GAPS')
        df_master = pd.read_excel(excel_file, sheet_name='Master')
    except FileNotFoundError:
        print(f"❌ Error: File not found: {excel_file}")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        sys.exit(1)
    
    print(f"✅ Loaded {len(df_gaps)} rows from GAPS sheet")
    print(f"✅ Loaded {len(df_master)} rows from Master sheet\n")
    
    # Get all STIs
    stis_from_gaps = set(df_gaps['STI'].dropna().unique())
    stis_from_master = set(get_all_stis_from_master(df_master))
    all_stis = sorted(list(stis_from_gaps | stis_from_master))
    
    print(f"✅ Found {len(all_stis)} unique STIs\n")
    
    # Process gaps from GAPS sheet
    gaps = []
    
    for idx, row in df_gaps.iterrows():
        gap_desc = str(row.get('Gap Description', '')).strip() if pd.notna(row.get('Gap Description')) else ''
        
        # Skip empty gap descriptions
        if not gap_desc or gap_desc.lower() in ['nan', 'none', '']:
            continue
        
        # Exclude "missing responses" and "missing documentation"
        gap_desc_lower = gap_desc.lower()
        if 'missing response' in gap_desc_lower or 'missing documentation' in gap_desc_lower:
            continue
        
        sti = str(row.get('STI', 'N/A')).strip() if pd.notna(row.get('STI')) else 'N/A'
        row_ref = str(row.get('Row Reference', '')).strip() if pd.notna(row.get('Row Reference')) else ''
        
        # Try to find corresponding row in Master sheet for more details
        master_details = {}
        if row_ref and 'Row' in row_ref:
            try:
                # Extract row number from "Row 36" format
                row_num_match = re.search(r'Row\s+(\d+)', row_ref)
                if row_num_match:
                    master_row_num = int(row_num_match.group(1)) - 1  # Convert to 0-indexed
                    if 0 <= master_row_num < len(df_master):
                        master_row = df_master.iloc[master_row_num]
                        master_details = {
                            'Question/Evidence Requirement': str(master_row.get('Questions/Evidence Requirements', '')).strip()[:500] if pd.notna(master_row.get('Questions/Evidence Requirements')) else '',
                            'BU Response Full': str(master_row.get('BU Response ', '')).strip()[:2000] if pd.notna(master_row.get('BU Response ')) else '',
                            'BU Answer': str(master_row.get('BU Answer', '')).strip()[:1000] if pd.notna(master_row.get('BU Answer')) else '',
                            'BU Supporting Evidence': str(master_row.get('BU Supporting Evidence', '')).strip()[:1000] if pd.notna(master_row.get('BU Supporting Evidence')) else '',
                            'Hints/Guidance': str(master_row.get('Hints/Guidance', '')).strip()[:500] if pd.notna(master_row.get('Hints/Guidance')) else '',
                            'Category': str(master_row.get('Category (Control)', '')).strip() if pd.notna(master_row.get('Category (Control)')) else '',
                        }
            except:
                pass
        
        # This is a gap
        gap_data = {
            'Gap Number': str(row.get('Gap Number', '')).strip() if pd.notna(row.get('Gap Number')) else idx + 1,
            'STI': sti,
            'Domain': str(row.get('Domain', 'N/A')).strip() if pd.notna(row.get('Domain')) else 'N/A',
            'ESS Control': str(row.get('ESS Control', 'N/A')).strip() if pd.notna(row.get('ESS Control')) and str(row.get('ESS Control')) != 'nan' else 'N/A',
            'AI Question': str(row.get('AI Question', 'N/A')).strip() if pd.notna(row.get('AI Question')) else 'N/A',
            'Gap Description': gap_desc,
            'Question/Evidence Requirement': master_details.get('Question/Evidence Requirement', ''),
            'BU Response Full': master_details.get('BU Response Full', ''),
            'BU Answer': master_details.get('BU Answer', ''),
            'BU Supporting Evidence': master_details.get('BU Supporting Evidence', ''),
            'Hints/Guidance': master_details.get('Hints/Guidance', ''),
            'Category': master_details.get('Category', ''),
            'Row Reference': row_ref
        }
        
        gaps.append(gap_data)
    
    print(f"✅ Found {len(gaps)} compliance gaps with details\n")
    
    # Create a comprehensive list with all STIs
    # For STIs without gaps, create entries showing they have no gaps
    stis_with_gaps = set([g['STI'] for g in gaps])
    stis_without_gaps = [sti for sti in all_stis if sti not in stis_with_gaps]
    
    # Add entries for STIs without gaps
    for sti in stis_without_gaps:
        gaps.append({
            'Gap Number': '',
            'STI': sti,
            'Domain': 'No Gaps',
            'ESS Control': '',
            'AI Question': '',
            'Gap Description': 'No compliance gaps identified',
            'Question/Evidence Requirement': '',
            'BU Response Full': '',
            'BU Answer': '',
            'BU Supporting Evidence': '',
            'Hints/Guidance': '',
            'Category': '',
            'Row Reference': ''
        })
    
    print(f"✅ Total entries: {len(gaps)} ({len(stis_with_gaps)} STIs with gaps, {len(stis_without_gaps)} STIs without gaps)\n")
    
    if not gaps:
        print("⚠️  No gaps found.")
        return
    
    # Create DataFrame
    df_output = pd.DataFrame(gaps)
    
    # Sort by STI, then by Gap Number
    df_output = df_output.sort_values(['STI', 'Gap Number'], na_position='last')
    
    # Create Excel file with formatting
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Sheet 1: All Gaps with Details
        df_output.to_excel(writer, sheet_name='All Gaps with Details', index=False)
        
        # Sheet 2: Summary by STI
        sti_summary = []
        for sti in all_stis:
            sti_gaps = [g for g in gaps if g['STI'] == sti and g['Domain'] != 'No Gaps']
            sti_summary.append({
                'STI': sti,
                'Total Gaps': len(sti_gaps),
                'Domains Affected': ', '.join(set([g['Domain'] for g in sti_gaps])) if sti_gaps else 'None',
                'ESS Controls Affected': ', '.join(set([g['ESS Control'] for g in sti_gaps if g['ESS Control'] != 'N/A'])) if sti_gaps else 'None',
                'Status': 'Has Gaps' if sti_gaps else 'No Gaps'
            })
        
        df_summary = pd.DataFrame(sti_summary)
        df_summary.to_excel(writer, sheet_name='STI Summary', index=False)
        
        # Format both sheets
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            
            # Style the header row
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
            
            # Auto-adjust column widths
            for col_idx, col in enumerate(df_output.columns if sheet_name == 'All Gaps with Details' else df_summary.columns, 1):
                max_length = 0
                col_letter = get_column_letter(col_idx)
                
                # Check header
                max_length = max(max_length, len(str(col)))
                
                # Check data rows
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    if row[col_idx - 1].value:
                        cell_value = str(row[col_idx - 1].value)
                        max_length = max(max_length, len(cell_value))
                
                adjusted_width = min(max_length + 2, 150)  # Cap at 150
                worksheet.column_dimensions[col_letter].width = adjusted_width
            
            # Wrap text in cells and add borders
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    cell.border = border
            
            # Freeze header row
            worksheet.freeze_panes = 'A2'
    
    print(f"✅ Excel file created: {output_file}")
    print(f"   Sheet 1: All Gaps with Details - {len(df_output)} rows")
    print(f"   Sheet 2: STI Summary - {len(df_summary)} STIs")
    print(f"\nGap breakdown by Domain:")
    domain_counts = df_output[df_output['Domain'] != 'No Gaps']['Domain'].value_counts()
    for domain, count in domain_counts.items():
        print(f"  {domain}: {count}")
    
    print(f"\nSTIs with most gaps:")
    sti_counts = df_output[df_output['Domain'] != 'No Gaps']['STI'].value_counts()
    for sti, count in sti_counts.head(10).items():
        print(f"  {sti}: {count} gaps")

def main():
    parser = argparse.ArgumentParser(description="Extract all Secure Flow SP compliance gaps with full details")
    parser.add_argument("--input", 
                       default="/Users/kwhelan/Downloads/AU2 Secure Flow + SSC - Simplified Questions - Internal Audit Prep - Template (4).xlsx",
                       help="Input Excel file path")
    parser.add_argument("--output", 
                       default="Secure_Flow_SP_Gaps_Complete.xlsx",
                       help="Output Excel file name")
    args = parser.parse_args()
    
    extract_gaps_with_details(args.input, args.output)

if __name__ == "__main__":
    main()




